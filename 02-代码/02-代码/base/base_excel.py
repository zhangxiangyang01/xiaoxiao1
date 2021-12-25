import openpyxl

excel_url = "./data/接口case.xlsx"


def load_excel():
    """
    加载excel
    """
    open_excel = openpyxl.load_workbook(excel_url)
    return open_excel


def get_sheet_data(index=None):
    """
    加载所有sheet的内容
    """
    sheet_name = load_excel().sheetnames
    if index is None:
        index = 0
    data = load_excel()[sheet_name[index]]
    return data


def get_cell_value(row, cols):
    """
    获取某一个单元格内容
    """
    data = get_sheet_data().cell(row=row, column=cols).value
    return data


def get_rows():
    """
    获取行数
    """
    row = get_sheet_data().max_row
    return row


def get_columns():
    """
    获取列数
    """
    column = get_sheet_data().max_column
    return column


def get_rows_value(row):
    """
    获取某一行的内容
    """
    row_list = []
    for i in get_sheet_data()[row]:
        row_list.append(i.value)
    return row_list


def get_columns_value(key):
    """
    获取某一列得数据
    """
    columns_list = []
    if not key.isalpha():
        raise Exception("key参数应为 'A/B/C' 这样的列名")
    key = key.upper()
    columns_list_data = get_sheet_data()[key]
    for i in columns_list_data:
        columns_list.append(i.value)
    return columns_list


def get_excel_data():
    """
    获取excel里面所有的数据
    """
    data_list = []
    for i in range(get_rows()):
        data_list.append(get_rows_value(i + 2))

    return data_list


def get_rows_number_with_case_id(case_id):
    """
    获取行号
    """
    num = 1
    cols_data = get_columns_value("B")
    for col_data in cols_data:
        if case_id == col_data:
            return num
        num = num + 1
    return num


def excel_write_data(row, cols, value):
    """
    写入数据
    """
    wb = load_excel()
    wr = wb.active
    wr.cell(row, cols, value)
    wb.save(excel_url)


def get_item_with_case_id(case_id):
    return ExcelModule(case_id)


class ExcelModule:

    def __init__(self, case_id):
        row = get_rows_number_with_case_id(case_id)
        values = get_rows_value(row)

        self.module_name = values[0]
        self.case_id = values[1]
        self.title = values[2]
        self.fixture = values[3]
        self.url = values[4]
        self.method = values[5]
        self.data = values[6]
        self.cookies = values[7]
        self.headers = values[8]
        self.expect_way = values[9]
        self.expect_res = values[10]
        self.desc = values[11]